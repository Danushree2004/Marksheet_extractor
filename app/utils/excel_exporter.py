from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from typing import Dict, Any, Optional, List

def get_value(field):
    """Extract value from field with confidence structure"""
    if field and isinstance(field, dict) and field.get("value") is not None:
        return field["value"]
    return ""

def export_to_excel(extracted_data: Dict[str, Any]) -> BytesIO:
    """
    Export single extraction to Excel.
    """
    return export_multiple_to_excel([extracted_data])

def export_multiple_to_excel(extractions_list: List[Dict[str, Any]]) -> BytesIO:
    """
    Convert multiple extracted marksheet/exam data to Excel format with all rows in one file.
    Supports batch export of multiple students.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Marksheet Data"
        
        # Define column headers - matching the exact user structure
        headers = [
            "#",
            "Roll No",
            "Name",
            "Batch",
            "Attendance",
            # Part A Questions (Q1-Q10, 2 marks each)
            "Q1 (2)", "Q2 (2)", "Q3 (2)", "Q4 (2)", "Q5 (2)",
            "Q6 (2)", "Q7 (2)", "Q8 (2)", "Q9 (2)", "Q10 (2)",
            # Part A Total
            "PART A Total Mark",
            # Part B Questions (Q11-Q14, 10 marks each)
            "Q11 (10)", "Q12 (10)", "Q13 (10)", "Q14 (10)",
            # Part B Total
            "PART B Total Mark",
            # Final totals
            "Total Mark", "%", "Status"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Process each extraction and add as a row
        for row_idx, extracted_data in enumerate(extractions_list, 2):
            _add_extraction_row(ws, extracted_data, row_idx)
        
        # Adjust column widths to match new header structure
        column_widths = {
            'A': 5,    # #
            'B': 12,   # Roll No
            'C': 15,   # Name
            'D': 12,   # Batch
            'E': 12,   # Attendance
            'F': 8, 'G': 8, 'H': 8, 'I': 8, 'J': 8,  # Q1-Q5 (Part A)
            'K': 8, 'L': 8, 'M': 8, 'N': 8, 'O': 8,  # Q6-Q10 (Part A)
            'P': 15,   # PART A Total Mark
            'Q': 10, 'R': 10, 'S': 10, 'T': 10,  # Q11-Q14 (Part B)
            'U': 15,   # PART B Total Mark
            'V': 12,   # Total Mark
            'W': 8,    # %
            'X': 10    # Status
        }
        
        from openpyxl.utils import get_column_letter
        for col_num, width in enumerate(column_widths.values(), 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
        
    except Exception as e:
        print(f"[EXCEL ERROR] Failed to create Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


def _add_extraction_row(ws, extracted_data: Dict[str, Any], row_idx: int):
    """
    Add a single extraction as a row in the worksheet.
    
    Args:
        ws: Openpyxl worksheet
        extracted_data: Extracted marksheet data
        row_idx: Row index to write data to
    """
    # Extract data
    candidate_details = extracted_data.get("candidate_details", {})
    exam_marks = extracted_data.get("exam_marks", {})
    exam_totals = extracted_data.get("exam_totals", {})
    
    # Initialize data row with row number
    data_row = [row_idx - 1]  # Row number starts at 1, but we count data rows
    
    # Get Roll Number
    roll_number = ""
    for field_name in ["roll_number", "register_number", "registration_number"]:
        field_data = candidate_details.get(field_name)
        if field_data:
            roll_number = get_value(field_data)
            if roll_number:
                break
    data_row.append(roll_number)
    
    # Get Name
    name = ""
    name_field = candidate_details.get("name")
    if name_field:
        name = get_value(name_field)
    data_row.append(name)
    
    # Get Batch (using programme or institution as fallback)
    batch = ""
    batch_field = candidate_details.get("programme")
    if batch_field:
        batch = get_value(batch_field)
    data_row.append(batch)
    
    # Attendance (not extracted, will be empty)
    data_row.append("")
    
    # Extract Part A questions (Q1-Q10, 2 marks each)
    part_a = exam_marks.get("part_a", {}) if exam_marks else {}
    part_a_questions = part_a.get("questions", []) if part_a else []
    
    for i in range(10):  # Q1 to Q10
        if i < len(part_a_questions):
            try:
                obtained = get_value(part_a_questions[i].get("obtained_marks"))
                data_row.append(obtained if obtained else "")
            except (KeyError, TypeError, AttributeError):
                data_row.append("")
        else:
            data_row.append("")
    
    # Get Part A Total
    part_a_total = ""
    if exam_totals:
        part_a_total = get_value(exam_totals.get("part_a_total"))
    data_row.append(part_a_total)
    
    # Extract Part B questions (Q11-Q14, 10 marks each)
    part_b = exam_marks.get("part_b", {}) if exam_marks else {}
    part_b_questions = part_b.get("questions", []) if part_b else []
    
    for i in range(4):  # Q11-Q14 (indices 0-3 in part_b_questions)
        if i < len(part_b_questions):
            try:
                obtained = get_value(part_b_questions[i].get("obtained_marks"))
                data_row.append(obtained if obtained else "")
            except (KeyError, TypeError, AttributeError):
                data_row.append("")
        else:
            data_row.append("")
    
    # Get Part B Total
    part_b_total = ""
    if exam_totals:
        part_b_total = get_value(exam_totals.get("part_b_total"))
    data_row.append(part_b_total)
    
    # Get Grand Total
    grand_total = ""
    max_marks = "50"
    percentage_str = ""
    
    if exam_totals:
        grand_total = get_value(exam_totals.get("grand_total"))
        max_marks = get_value(exam_totals.get("max_marks")) or "50"
    
        # Calculate percentage
        try:
            grand_total_val = float(grand_total) if grand_total else 0
            max_marks_val = float(max_marks) if max_marks else 50
            percentage = (grand_total_val / max_marks_val * 100) if max_marks_val > 0 else 0
            percentage_str = f"{percentage:.2f}"
        except (ValueError, ZeroDivisionError, TypeError):
            percentage_str = ""
    
    data_row.append(grand_total)
    data_row.append(percentage_str)
    
    # Status (Pass/Fail based on percentage, or empty if not enough data)
    status = ""
    if percentage_str:
        try:
            perc_val = float(percentage_str)
            status = "Pass" if perc_val >= 40 else "Fail"  # Typical passing percentage
        except ValueError:
            pass
    data_row.append(status)
    
    # Write data row
    for col_idx, value in enumerate(data_row, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
